import openpyxl

#data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb=openpyxl.load_workbook("C:/Users/nye/PycharmProjects/chatbot/test.xlsx", data_only=True)

#시트이름 불러오기
load_ws=load_wb['2020']

#셀 주소로 값 출력
print(load_ws['F4'].value)

#셀 좌표로 값 출력
print(load_ws.cell(6,4).value)


print('\n-----지정한 셀 출력-----')
get_cells = load_ws['A2':'F6']
for row in get_cells:
        for cell in row:
            print(cell.value)



print('\n-----모든 행 단위로 출력-----')
for row in load_ws.rows:
    print(row)


print('\n-----모든 열 단위로 출력-----')
for column in load_ws.columns:
    print(column)


print('\n-----모든 행과 열 출력-----')
all_values = []
for row in load_ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)




write_wb = openpyxl.Workbook()


# 이름이 있는 시트를 생성
ws2 = write_wb.create_sheet('시트예제')

# Sheet1에다 입력
write_ws = write_wb.active
write_ws['A1'] = '숫자'

# 행 단위로 추가
write_ws.append([1, 2, 3])
write_ws.append([4, 5, 6])

ws2['A2']='두번째 시트'

# 셀 단위로 추가
write_ws.cell(5, 5, '5행5열')
write_wb.save("C:/Users/nye/PycharmProjects/chatbot/test2.xlsx")


